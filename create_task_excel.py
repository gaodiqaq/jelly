import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# 创建目录
os.makedirs(r"D:\gold_work", exist_ok=True)

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "个人任务台账"

# ==================== 设置样式 ====================
# 标题样式
title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# 表头样式
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 数据单元格样式
data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
data_alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 边框
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# 条件格式填充色
done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 已完成 - 绿色
todo_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 待办 - 黄色
progress_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")  # 进行中 - 蓝色
cancel_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # 已取消 - 橙色

high_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # 高 - 红色
medium_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # 中 - 橙色
low_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 低 - 绿色

# ==================== 写入标题 ====================
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "个人任务台账"
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# ==================== 写入表头 ====================
headers = [
    ("编号", 8),
    ("任务名称", 25),
    ("优先级", 10),
    ("状态", 10),
    ("创建日期", 12),
    ("截止日期", 12),
    ("分类", 10),
    ("备注", 30)
]

for col_idx, (header, width) in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[2].height = 25

# ==================== 设置下拉菜单 ====================
# 优先级下拉
priority_dv = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
priority_dv.error = "请选择优先级：高、中、低"
priority_dv.errorTitle = "输入错误"
priority_dv.prompt = "请选择优先级"
priority_dv.promptTitle = "优先级"
ws.add_data_validation(priority_dv)
priority_dv.add(f"C3:C1000")  # 应用到 C3 到 C1000 行

# 状态下拉
status_dv = DataValidation(type="list", formula1='"待办,进行中,已完成,已取消"', allow_blank=True)
status_dv.error = "请选择状态：待办、进行中、已完成、已取消"
status_dv.errorTitle = "输入错误"
status_dv.prompt = "请选择状态"
status_dv.promptTitle = "状态"
ws.add_data_validation(status_dv)
status_dv.add(f"D3:D1000")

# 分类下拉
category_dv = DataValidation(type="list", formula1='"工作,学习,生活,健康,其他"', allow_blank=True)
category_dv.error = "请选择分类：工作、学习、生活、健康、其他"
category_dv.errorTitle = "输入错误"
category_dv.prompt = "请选择分类"
category_dv.promptTitle = "分类"
ws.add_data_validation(category_dv)
category_dv.add(f"G3:G1000")

# ==================== 条件格式 ====================
from openpyxl.formatting.rule import CellIsRule

# 状态条件格式
ws.conditional_formatting.add("D3:D1000",
    CellIsRule(operator="equal", formula=['"已完成"'], fill=done_fill)
)
ws.conditional_formatting.add("D3:D1000",
    CellIsRule(operator="equal", formula=['"待办"'], fill=todo_fill)
)
ws.conditional_formatting.add("D3:D1000",
    CellIsRule(operator="equal", formula=['"进行中"'], fill=progress_fill)
)
ws.conditional_formatting.add("D3:D1000",
    CellIsRule(operator="equal", formula=['"已取消"'], fill=cancel_fill)
)

# 优先级条件格式
ws.conditional_formatting.add("E3:E1000",
    CellIsRule(operator="equal", formula=['"高"'], fill=high_fill)
)
ws.conditional_formatting.add("E3:E1000",
    CellIsRule(operator="equal", formula=['"中"'], fill=medium_fill)
)
ws.conditional_formatting.add("E3:E1000",
    CellIsRule(operator="equal", formula=['"低"'], fill=low_fill)
)

# ==================== 设置数据格式 ====================
# 日期列格式
for row in range(3, 1001):
    ws.cell(row=row, column=5).number_format = "yyyy-mm-dd"  # 创建日期
    ws.cell(row=row, column=6).number_format = "yyyy-mm-dd"  # 截止日期

# ==================== 冻结窗格 ====================
ws.freeze_panes = "A3"  # 冻结第一行表头

# ==================== 自动筛选 ====================
ws.auto_filter.ref = f"A2:H2"

# ==================== 添加示例数据 ====================
sample_data = [
    ("T001", "完成项目周报", "高", "进行中", "2026-08-18", "2026-08-20", "工作", "周五前提交给领导"),
    ("T002", "购买生活用品", "中", "待办", "2026-08-18", "2026-08-22", "生活", "牙膏、纸巾、洗衣液"),
    ("T003", "学习 Python 基础", "低", "已完成", "2026-08-15", "2026-08-18", "学习", "完成基础语法学习"),
    ("T004", "预约体检", "中", "待办", "2026-08-18", "2026-08-25", "健康", "上午9点前到"),
    ("T005", "阅读《原子习惯》", "低", "进行中", "2026-08-17", "2026-08-30", "学习", "每天读30分钟"),
]

for row_idx, data in enumerate(sample_data, start=3):
    for col_idx, value in enumerate(data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [1, 3, 4, 5, 6, 7]:
            cell.alignment = data_alignment_center
        else:
            cell.alignment = data_alignment

# ==================== 保存文件 ====================
output_path = r"D:\gold_work\个人任务台账.xlsx"
wb.save(output_path)
print("Excel 文件已创建：" + output_path)
print("\n功能说明：")
print("  1. 优先级列（C列）：点击下拉选择【高 / 中 / 低】")
print("  2. 状态列（D列）：点击下拉选择【待办 / 进行中 / 已完成 / 已取消】")
print("  3. 分类列（G列）：点击下拉选择【工作 / 学习 / 生活 / 健康 / 其他】")
print("  4. 条件格式：状态和优先级会自动着色")
print("  5. 已开启自动筛选和冻结表头")
print("  6. 已预填 5 条示例数据供参考")